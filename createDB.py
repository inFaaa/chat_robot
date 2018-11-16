from sqlalchemy import create_engine,MetaData
from sqlalchemy.orm import sessionmaker
from app.models.base import Base,db
from app.models.query_pair import QueryPair
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.declarative import declarative_base



wb = load_workbook(filename="./testfile.xlsx")
#第一张表
sheet_first = wb.get_sheet_by_name(wb.get_sheet_names()[0])
rows = sheet_first.rows

insert_data = []
for row in rows:
    pair = [col.value for col in row]
    print(pair)
    insert_data.append(pair)



# test_data =xlrd.open_workbook("./testfile.xlsx") .sheet_by_index(0)

engine = create_engine("mysql+pymysql://root:123456@localhost:3306/querypair")

# Base.metadata.create_all(engine)

DBSession = sessionmaker(bind=engine)
session = DBSession()

#测试用
# qp = QueryPair()
# qp.query_string = insert_data[1][0]
# qp.answer = insert_data[1][1]
# session.add(qp)
# session.commit()

#全插
for i in range(len(insert_data)):
    qp = QueryPair()
    qp.query_string = insert_data[i][0]
    qp.answer = insert_data[i][1]
    session.add(qp)
    session.commit()

# qp1 = session.query(QueryPair).filter(QueryPair.query_string=="进程").all()
# print(qp1.query_string)

session.close()
