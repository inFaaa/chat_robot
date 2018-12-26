from openpyxl import Workbook,load_workbook
import jieba.posseg as posg

PATH = "./knowledge-export-1.xlsx"
wb = load_workbook(filename=PATH)
sheet_first = wb.get_sheet_by_name(wb.get_sheet_names()[0])
rows = sheet_first.rows

new_wb = Workbook()
new_sheet = new_wb.active

word_list = []
for row in rows:
    for item in posg.cut(str(row[1].value)):
        if(item.flag=="n"):
            word_list.append(item.word)
    for item in posg.cut(str(row[2].value)):
        if(item.flag=="n"):
            word_list.append(item.word)

word_list = list(set(word_list))#去重
for i in range(len(word_list)):
    new_sheet.append([word_list[i]])

new_wb.save("./new_sheet.xlsx")